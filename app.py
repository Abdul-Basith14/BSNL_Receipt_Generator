"""
CASH RECEIPTS GENERATOR - WEB APPLICATION
==========================================
A web interface for generating cash receipts from Excel data.

Features:
- Upload Excel file with TY Adv Appl format
- Automatic receipt generation
- Download generated receipts
- Deployable web application

Usage:
    python app.py
    Open browser to http://localhost:5000
"""

from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import re
import os
from werkzeug.utils import secure_filename
import io
import tempfile
import uuid
import random

app = Flask(__name__)
app.secret_key = 'bsnl_cash_receipts_secret_key_2026'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Contractor pools for PITS and OH Cable work
CONTRACTORS_PITS = [
    "Tilak G, 7th Cross, Veerasagara, Tumkur",
    "K G Ravi Kaidala Tumkur",
    "Narasimha Murthy, Kittadakuppe, Gubbi",
    "Siddappa, Kaidala, Gulur Hobli, Tumkur"
]

CONTRACTORS_OH_CABLE = [
    "Tilak G, 7th Cross, Veerasagara, Tumkur",
    "K G Ravi Kaidala Tumkur",
    "Narasimha Murthy, Kittadakuppe, Gubbi",
    "Siddappa, Kaidala, Gulur Hobli, Tumkur"
]

# Create directories if they don't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def number_to_words(num):
    """Convert number to words in Indian numbering system"""
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", 
             "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    
    if num == 0:
        return "Zero"
    
    def convert_hundreds(n):
        if n == 0:
            return ""
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
        else:
            return ones[n // 100] + " Hundred" + (" " + convert_hundreds(n % 100) if n % 100 != 0 else "")
    
    if num < 1000:
        return convert_hundreds(num)
    elif num < 100000:
        thousands = num // 1000
        remainder = num % 1000
        result = convert_hundreds(thousands) + " Thousand"
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result
    elif num < 10000000:
        lakhs = num // 100000
        remainder = num % 100000
        result = convert_hundreds(lakhs) + " Lakh"
        if remainder >= 1000:
            result += " " + convert_hundreds(remainder // 1000) + " Thousand"
            remainder = remainder % 1000
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result
    else:
        crores = num // 10000000
        remainder = num % 10000000
        result = convert_hundreds(crores) + " Crore"
        if remainder >= 100000:
            result += " " + convert_hundreds(remainder // 100000) + " Lakh"
            remainder = remainder % 100000
        if remainder >= 1000:
            result += " " + convert_hundreds(remainder // 1000) + " Thousand"
            remainder = remainder % 1000
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result

def generate_description_pits(date_obj, work_details, route, amount, contractor_name):
    """Generate description for Pits work"""
    # Ensure date_obj is datetime
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
        except:
            try:
                date_obj = datetime.strptime(date_obj, '%d-%m-%Y')
            except:
                date_obj = datetime.now()
    
    pits_match = re.search(r'(\d+)\s*pits?', work_details, re.IGNORECASE)
    num_pits = pits_match.group(1) if pits_match else "2"
    
    location_info = ""
    dist_match = re.search(r'(\d+\.\d+)km', work_details, re.IGNORECASE)
    if dist_match:
        distance = dist_match.group(1)
        from_match = re.search(r'from\s+([A-Za-z\s]+?)(?:\s+in|\s+due|\s+,|\.)', work_details, re.IGNORECASE)
        from_location = from_match.group(1).strip() if from_match else route.split()[0] if route else "Exchange"
        location_info = f"at OTDR Distance {distance}Km from {from_location}"
    else:
        location_match = re.search(r'at\s+([^,\.]+)', work_details, re.IGNORECASE)
        if location_match:
            location_info = f"at {location_match.group(1).strip()}"
        else:
            location_info = f"on {route}"
    
    route_desc = route if route else "OFC route"
    
    reason = ""
    if "water" in work_details.lower() or "pipeline" in work_details.lower():
        reason = "due to JJM water pipeline trenching work"
    elif "road" in work_details.lower() or "nh" in work_details.lower():
        reason = "due to road work"
    elif "bescom" in work_details.lower():
        reason = "due to BESCOM work"
    elif "rly" in work_details.lower() or "railway" in work_details.lower():
        reason = "due to Railway work"
    else:
        reason = "for fault restoration"
    
    date_str = date_obj.strftime("%d-%m-%Y")
    
    description = (f"Paid Charges to {contractor_name}  and Team "
                  f"Rs {amount} Towards opening {num_pits} nos of Joint pits Pits  and "
                  f"Trenching between the joints {location_info} for attending "
                  f"OFC cable cut on {route_desc} {reason} and closed the opened pits "
                  f"by backfilling the excavated trenchs after  restoration of fault on "
                  f"{date_str}. the work is carried out On contract basis.")
    
    return description

def generate_description_oh_cable(date_obj, work_details, route, amount, contractor_name):
    """Generate description for OH Cable work"""
    # Ensure date_obj is datetime
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
        except:
            try:
                date_obj = datetime.strptime(date_obj, '%d-%m-%Y')
            except:
                date_obj = datetime.now()
    
    length_match = re.search(r'(\d+)\s*mtr', work_details, re.IGNORECASE)
    length = length_match.group(1) if length_match else "100"
    
    location_info = ""
    dist_match = re.search(r'(\d+\.\d+)km', work_details, re.IGNORECASE)
    if dist_match:
        distance = dist_match.group(1)
        from_match = re.search(r'from\s+([A-Za-z\s]+?)(?:\s+in|\s+due|\.|\,)', work_details, re.IGNORECASE)
        from_location = from_match.group(1).strip() if from_match else route.split()[0] if route else "Exchange"
        location_info = f"at OTDR Distance {distance}Km from {from_location}"
    elif "at" in work_details.lower():
        at_match = re.search(r'at\s+([^,\.]+)', work_details, re.IGNORECASE)
        if at_match:
            location_info = f"at {at_match.group(1).strip()}"
    else:
        location_info = f"on {route}"
    
    route_desc = route if route else "OFC route"
    
    reason = ""
    if "bescom" in work_details.lower():
        reason = "due to BESCOM work"
    elif "road" in work_details.lower():
        reason = "due to road work"
    elif "monkey" in work_details.lower():
        reason = "due to Monkey bite"
    elif "water" in work_details.lower() or "pipeline" in work_details.lower():
        reason = "due to water pipeline work"
    else:
        reason = "for fault restoration"
    
    date_str = date_obj.strftime("%d-%m-%Y")
    
    description = (f"Paid Charges to {contractor_name}  and Team "
                  f"Rs {amount} Towards layed {length}Mtr OH cable for attending "
                  f"OH cable cut {location_info} on {route_desc} {reason} "
                  f"and restoration of fault on {date_str}. "
                  f"the work is carried out On contract basis.")
    
    return description

def generate_receipts(input_file):
    """Generate cash receipts from uploaded Excel file"""
    # Load workbook
    wb_source = openpyxl.load_workbook(input_file)
    
    # Try to find TY Adv Appl sheet or similar
    sheet_name = None
    for name in wb_source.sheetnames:
        if 'ty' in name.lower() and 'adv' in name.lower():
            sheet_name = name
            break
    
    if not sheet_name:
        sheet_name = wb_source.sheetnames[0]  # Use first sheet
    
    ws_ty = wb_source[sheet_name]
    
    # Read data
    ty_data = []
    for row_num in range(4, 100):  # Check up to row 100
        date_val = ws_ty.cell(row_num, 1).value
        route = ws_ty.cell(row_num, 2).value
        work_details = ws_ty.cell(row_num, 3).value
        pits_oh = ws_ty.cell(row_num, 7).value
        amount = ws_ty.cell(row_num, 8).value
        
        if not date_val or not amount or work_details in ["Local Purchase", "Total", None]:
            continue
        
        if isinstance(date_val, datetime):
            date_obj = date_val
        else:
            try:
                date_obj = datetime.strptime(str(date_val), "%Y-%m-%d %H:%M:%S")
            except:
                try:
                    date_obj = datetime.strptime(str(date_val), "%Y-%m-%d")
                except:
                    continue
        
        ty_data.append({
            'date': date_obj,
            'route': route if route else "",
            'work_details': work_details if work_details else "",
            'pits_oh': pits_oh if pits_oh else "",
            'amount': int(amount) if amount else 0
        })
    
    if not ty_data:
        return None, "No valid data found in the uploaded file"
    
    # Create new workbook
    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Cash Receipts"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 12
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    voucher_no = 1
    
    # Track contractor assignments per day to ensure variety
    contractor_assignments = {}
    
    # Process each entry
    for entry in ty_data:
        pits_oh_lower = entry['pits_oh'].lower()
        is_pits = 'pit' in pits_oh_lower
        
        # Ensure date is datetime object
        date_obj = entry['date']
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            except:
                try:
                    date_obj = datetime.strptime(date_obj, '%d-%m-%Y')
                except:
                    date_obj = datetime.now()
        
        # Select contractor based on work type and date
        date_key = date_obj.strftime('%Y-%m-%d')
        
        # Get appropriate contractor list
        contractor_pool = CONTRACTORS_PITS if is_pits else CONTRACTORS_OH_CABLE
        
        # Try to assign different contractor per day
        if date_key in contractor_assignments:
            # Exclude already used contractor for this date
            available = [c for c in contractor_pool if c != contractor_assignments.get(date_key)]
            contractor_name = random.choice(available) if available else random.choice(contractor_pool)
        else:
            contractor_name = random.choice(contractor_pool)
        
        # Store assignment
        contractor_assignments[date_key] = contractor_name
        
        if is_pits:
            description = generate_description_pits(date_obj, entry['work_details'], 
                                                   entry['route'], entry['amount'], contractor_name)
        else:
            description = generate_description_oh_cable(date_obj, entry['work_details'], 
                                                        entry['route'], entry['amount'], contractor_name)
        
        amount_words = number_to_words(entry['amount'])
        
        # Row 1: CASH RECEIPT
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "CASH RECEIPT"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
        
        # Row 2: Date and Voucher No
        ws[f'A{current_row}'] = "Date"
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws[f'B{current_row}'] = date_obj
        ws[f'B{current_row}'].number_format = 'DD-MM-YYYY'
        ws[f'B{current_row}'].border = thin_border
        ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'C{current_row}:F{current_row}')
        for col in ['C', 'D', 'E', 'F']:
            ws[f'{col}{current_row}'].border = thin_border
        
        ws[f'G{current_row}'] = "Voucher No:"
        ws[f'G{current_row}'].border = thin_border
        ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'H{current_row}'] = voucher_no
        ws[f'H{current_row}'].border = thin_border
        ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Row 3: Received from
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Received from SDE (Txn), Tumkur  Sum of Rupees {entry['amount']}/-"
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
        
        # Row 4: Description
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = description
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.row_dimensions[current_row].height = 75
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
        
        # Row 5: RS and Amount
        ws[f'A{current_row}'] = "RS"
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells(f'B{current_row}:H{current_row}')
        cell = ws[f'B{current_row}']
        cell.value = entry['amount']
        cell.alignment = Alignment(horizontal='left', vertical='center')
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
        
        # Row 6: Amount in words + Signatures
        ws.merge_cells(f'A{current_row}:D{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"Rupees {amount_words} only"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        
        ws.merge_cells(f'E{current_row}:F{current_row}')
        cell = ws[f'E{current_row}']
        cell.value = "Signature of Payee"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['E', 'F']:
            ws[f'{col}{current_row}'].border = thin_border
        
        ws.merge_cells(f'G{current_row}:H{current_row}')
        cell = ws[f'G{current_row}']
        cell.value = "Signature of witness"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in ['G', 'H']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
        
        # Rows 7-12: Standard clauses
        for text in [
            "1. Labour Engaged is Justified",
            "2.Work is done satisfactorily",
            "3.Provision Exists in the estimate Maintainnace Grant",
            "RM Cables / TMR/LABOUR/5020819",
            f"Passed and Paid for Rs. {entry['amount']}/-",
            f"(Rupees {amount_words.title()} only)"
        ]:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = text
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col}{current_row}'].border = thin_border
            current_row += 1
        
        current_row += 2
        voucher_no += 1
    
    wb_source.close()
    
    # Prepare preview data (all receipts with full details)
    preview_data = []
    contractor_preview_assignments = {}
    voucher_no_preview = 1
    
    for i, entry in enumerate(ty_data):
        pits_oh_lower = entry['pits_oh'].lower()
        is_pits = 'pit' in pits_oh_lower
        work_type = "PITS Work" if is_pits else "OH Cable Work"
        
        # Handle date formatting safely
        date_obj = entry['date']
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d')
            except:
                try:
                    date_obj = datetime.strptime(date_obj, '%d-%m-%Y')
                except:
                    date_obj = datetime.now()
        
        # Select contractor for preview (same logic as main generation)
        date_key = date_obj.strftime('%Y-%m-%d')
        work_key = f"{date_key}_{work_type}"
        
        if work_key in contractor_preview_assignments:
            selected_contractor = contractor_preview_assignments[work_key]
        else:
            contractor_pool = CONTRACTORS_PITS if is_pits else CONTRACTORS_OH_CABLE
            date_contractors = [v for k, v in contractor_preview_assignments.items() if k.startswith(date_key)]
            available = [c for c in contractor_pool if c not in date_contractors]
            if not available:
                available = contractor_pool
            selected_contractor = random.choice(available)
            contractor_preview_assignments[work_key] = selected_contractor
        
        # Generate description
        date_formatted = date_obj.strftime('%d-%m-%Y')
        if is_pits:
            description = generate_description_pits(
                date_formatted,
                entry['work_details'],
                entry['route'],
                entry['amount'],
                selected_contractor
            )
        else:
            description = generate_description_oh_cable(
                date_formatted,
                entry['work_details'],
                entry['route'],
                entry['amount'],
                selected_contractor
            )
        
        amount_words = number_to_words(entry['amount'])
        
        preview_data.append({
            'voucher_no': voucher_no_preview,
            'date': date_formatted,
            'work_type': work_type,
            'contractor': selected_contractor,
            'description': description,
            'amount': entry['amount'],
            'amount_words': amount_words,
            'route': entry['route']
        })
        voucher_no_preview += 1
    
    return wb_new, None, len(ty_data), preview_data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate():
    if 'file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        try:
            # Get original filename without extension
            original_name = os.path.splitext(file.filename)[0]
            
            wb_output, error, receipts_count, preview_data = generate_receipts(file)
            
            if error:
                flash(error, 'error')
                return redirect(url_for('index'))
            
            # Generate filename based on uploaded file name
            unique_id = str(uuid.uuid4())[:8]
            output_filename = f'{original_name}_cash_receipt_{unique_id}.xlsx'
            output_path = os.path.join(app.config['OUTPUT_FOLDER'], output_filename)
            
            # Save the workbook
            wb_output.save(output_path)
            wb_output.close()
            
            # Store data in session
            session['generated_file'] = output_filename
            session['receipts_count'] = receipts_count
            session['preview_data'] = preview_data
            session['original_filename'] = file.filename
            
            return redirect(url_for('preview'))
            
        except Exception as e:
            flash(f'Error processing file: {str(e)}', 'error')
            return redirect(url_for('index'))
    else:
        flash('Invalid file type. Please upload an Excel file (.xlsx or .xls)', 'error')
        return redirect(url_for('index'))

@app.route('/preview')
def preview():
    if 'generated_file' not in session:
        flash('No file to preview', 'error')
        return redirect(url_for('index'))
    
    return render_template('preview.html',
                         filename=session['generated_file'],
                         receipts_count=session.get('receipts_count', 0),
                         preview_data=session.get('preview_data', []),
                         original_filename=session.get('original_filename', 'Unknown'))

@app.route('/download/<filename>')
def download(filename):
    try:
        # Security check - ensure filename is in session
        if 'generated_file' not in session or session['generated_file'] != filename:
            flash('Invalid download request', 'error')
            return redirect(url_for('index'))
        
        file_path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
        
        if not os.path.exists(file_path):
            flash('File not found', 'error')
            return redirect(url_for('index'))
        
        # Use original filename for download
        download_name = session.get('generated_file', 'Cash_Receipts_Generated.xlsx')
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=download_name
        )
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('index'))

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
