"""
CASH RECEIPTS GENERATOR
=======================
This script generates formatted cash receipts from TY Adv Appl data.

Input:  Dec -25.xlsx (TY Adv Appl sheet)
Output: Final_Cash_Receipts.xlsx

Features:
- Generates cash receipts for each entry in TY Adv Appl
- Follows exact format with merged cells and borders
- Separate patterns for PITS work and OH Cable work
- Automatically converts amounts to Indian numbering words
- Includes all standard clauses and account codes

Usage:
    python generate_cash_receipts_final.py
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import re
import random

# Contractor pools for PITS and OH Cable work
CONTRACTORS_PITS = [
    "Tilak G, 7th Cross, Veerasagara, Tumkur",
    "K G Ravi Kaidala Tumkur",
    "Narasimha Murthy, Kittadakuppe, Gubbi",
    "Siddappa, Kaidala, Gulur Hobli, Tumkur"
]

CONTRACTORS_OH_CABLE = [
    "Tilak G, 7th Cross, Veerasagara, Tumkur",
    "K G Ravi Kaidala Tumkur",
    "Narasimha Murthy, Kittadakuppe, Gubbi",
    "Siddappa, Kaidala, Gulur Hobli, Tumkur"
]

def number_to_words(num):
    """Convert number to words in Indian numbering system"""
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", 
             "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    
    if num == 0:
        return "Zero"
    
    def convert_hundreds(n):
        if n == 0:
            return ""
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
        else:
            return ones[n // 100] + " Hundred" + (" " + convert_hundreds(n % 100) if n % 100 != 0 else "")
    
    if num < 1000:
        return convert_hundreds(num)
    elif num < 100000:
        thousands = num // 1000
        remainder = num % 1000
        result = convert_hundreds(thousands) + " Thousand"
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result
    elif num < 10000000:
        lakhs = num // 100000
        remainder = num % 100000
        result = convert_hundreds(lakhs) + " Lakh"
        if remainder >= 1000:
            result += " " + convert_hundreds(remainder // 1000) + " Thousand"
            remainder = remainder % 1000
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result
    else:
        crores = num // 10000000
        remainder = num % 10000000
        result = convert_hundreds(crores) + " Crore"
        if remainder >= 100000:
            result += " " + convert_hundreds(remainder // 100000) + " Lakh"
            remainder = remainder % 100000
        if remainder >= 1000:
            result += " " + convert_hundreds(remainder // 1000) + " Thousand"
            remainder = remainder % 1000
        if remainder:
            result += " " + convert_hundreds(remainder)
        return result

def generate_description_pits(date_obj, work_details, route, amount, contractor_name):
    """Generate description for Pits work"""
    pits_match = re.search(r'(\d+)\s*pits?', work_details, re.IGNORECASE)
    num_pits = pits_match.group(1) if pits_match else "2"
    
    location_info = ""
    dist_match = re.search(r'(\d+\.\d+)km', work_details, re.IGNORECASE)
    if dist_match:
        distance = dist_match.group(1)
        from_match = re.search(r'from\s+([A-Za-z\s]+?)(?:\s+in|\s+due|\s+,|\.)', work_details, re.IGNORECASE)
        from_location = from_match.group(1).strip() if from_match else route.split()[0] if route else "Exchange"
        location_info = f"at OTDR Distance {distance}Km from {from_location}"
    else:
        location_match = re.search(r'at\s+([^,\.]+)', work_details, re.IGNORECASE)
        if location_match:
            location_info = f"at {location_match.group(1).strip()}"
        else:
            location_info = f"on {route}"
    
    route_desc = route if route else "OFC route"
    
    reason = ""
    if "water" in work_details.lower() or "pipeline" in work_details.lower():
        reason = "due to JJM water pipeline trenching work"
    elif "road" in work_details.lower() or "nh" in work_details.lower():
        reason = "due to road work"
    elif "bescom" in work_details.lower():
        reason = "due to BESCOM work"
    elif "rly" in work_details.lower() or "railway" in work_details.lower():
        reason = "due to Railway work"
    else:
        reason = "for fault restoration"
    
    date_str = date_obj.strftime("%d-%m-%Y")
    
    description = (f"Paid Charges to {contractor_name}  and Team "
                  f"Rs {amount} Towards opening {num_pits} nos of Joint pits Pits  and "
                  f"Trenching between the joints {location_info} for attending "
                  f"OFC cable cut on {route_desc} {reason} and closed the opened pits "
                  f"by backfilling the excavated trenchs after  restoration of fault on "
                  f"{date_str}. the work is carried out On contract basis.")
    
    return description

def generate_description_oh_cable(date_obj, work_details, route, amount, contractor_name):
    """Generate description for OH Cable work"""
    length_match = re.search(r'(\d+)\s*mtr', work_details, re.IGNORECASE)
    length = length_match.group(1) if length_match else "100"
    
    location_info = ""
    dist_match = re.search(r'(\d+\.\d+)km', work_details, re.IGNORECASE)
    if dist_match:
        distance = dist_match.group(1)
        from_match = re.search(r'from\s+([A-Za-z\s]+?)(?:\s+in|\s+due|\.|\,)', work_details, re.IGNORECASE)
        from_location = from_match.group(1).strip() if from_match else route.split()[0] if route else "Exchange"
        location_info = f"at OTDR Distance {distance}Km from {from_location}"
    elif "at" in work_details.lower():
        at_match = re.search(r'at\s+([^,\.]+)', work_details, re.IGNORECASE)
        if at_match:
            location_info = f"at {at_match.group(1).strip()}"
    else:
        location_info = f"on {route}"
    
    route_desc = route if route else "OFC route"
    
    reason = ""
    if "bescom" in work_details.lower():
        reason = "due to BESCOM work"
    elif "road" in work_details.lower():
        reason = "due to road work"
    elif "monkey" in work_details.lower():
        reason = "due to Monkey bite"
    elif "water" in work_details.lower() or "pipeline" in work_details.lower():
        reason = "due to water pipeline work"
    else:
        reason = "for fault restoration"
    
    date_str = date_obj.strftime("%d-%m-%Y")
    
    description = (f"Paid Charges to {contractor_name}  and Team "
                  f"Rs {amount} Towards layed {length}Mtr OH cable for attending "
                  f"OH cable cut {location_info} on {route_desc} {reason} "
                  f"and restoration of fault on {date_str}. "
                  f"the work is carried out On contract basis.")
    
    return description

# Load source workbook
input_filename = 'Dec -25.xlsx'
print(f"Loading {input_filename}...")
wb_source = openpyxl.load_workbook(input_filename)
ws_ty = wb_source['TY Adv Appl']

# Extract base name for output file
import os
base_name = os.path.splitext(input_filename)[0]

# Read TY Adv Appl data
print("Reading TY Adv Appl data...")
ty_data = []
for row_num in range(4, 21):
    date_val = ws_ty.cell(row_num, 1).value
    route = ws_ty.cell(row_num, 2).value
    work_details = ws_ty.cell(row_num, 3).value
    pits_oh = ws_ty.cell(row_num, 7).value
    amount = ws_ty.cell(row_num, 8).value
    
    if not date_val or not amount or work_details in ["Local Purchase", "Total"]:
        continue
    
    if isinstance(date_val, datetime):
        date_obj = date_val
    else:
        try:
            date_obj = datetime.strptime(str(date_val), "%Y-%m-%d %H:%M:%S")
        except:
            date_obj = datetime.now()
    
    ty_data.append({
        'date': date_obj,
        'route': route if route else "",
        'work_details': work_details if work_details else "",
        'pits_oh': pits_oh if pits_oh else "",
        'amount': int(amount) if amount else 0
    })

print(f"Found {len(ty_data)} entries to process")

# Create new workbook
print("Creating formatted Cash Receipts workbook...")
wb_new = openpyxl.Workbook()
ws = wb_new.active
ws.title = "Cash Receipts"

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12

# Borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

current_row = 1
voucher_no = 1

# Track contractor assignments per day to ensure variety
contractor_assignments = {}

# Process each entry
for idx, entry in enumerate(ty_data, 1):
    print(f"Processing entry {idx}/{len(ty_data)}: {entry['date'].strftime('%Y-%m-%d')}, Amount: {entry['amount']}")
    
    pits_oh_lower = entry['pits_oh'].lower()
    is_pits = 'pit' in pits_oh_lower
    
    # Select contractor based on work type and date
    date_key = entry['date'].strftime('%Y-%m-%d')
    
    # Get appropriate contractor list
    contractor_pool = CONTRACTORS_PITS if is_pits else CONTRACTORS_OH_CABLE
    
    # Try to assign different contractor per day
    if date_key in contractor_assignments:
        # Exclude already used contractor for this date
        available = [c for c in contractor_pool if c != contractor_assignments.get(date_key)]
        contractor_name = random.choice(available) if available else random.choice(contractor_pool)
    else:
        contractor_name = random.choice(contractor_pool)
    
    # Store assignment
    contractor_assignments[date_key] = contractor_name
    
    if is_pits:
        description = generate_description_pits(entry['date'], entry['work_details'], 
                                               entry['route'], entry['amount'], contractor_name)
    else:
        description = generate_description_oh_cable(entry['date'], entry['work_details'], 
                                                    entry['route'], entry['amount'], contractor_name)
    
    amount_words = number_to_words(entry['amount'])
    
    # Row 1: CASH RECEIPT (merged across A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "CASH RECEIPT"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 2: Date and Voucher No
    ws[f'A{current_row}'] = "Date"
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'B{current_row}'] = entry['date']
    ws[f'B{current_row}'].number_format = 'DD-MM-YYYY'
    ws[f'B{current_row}'].border = thin_border
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Merge C-F for spacing
    ws.merge_cells(f'C{current_row}:F{current_row}')
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border
    
    ws[f'G{current_row}'] = "Voucher No:"
    ws[f'G{current_row}'].border = thin_border
    ws[f'G{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws[f'H{current_row}'] = voucher_no
    ws[f'H{current_row}'].border = thin_border
    ws[f'H{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    current_row += 1
    
    # Row 3: Received from (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"Received from SDE (Txn), Tumkur  Sum of Rupees {entry['amount']}/-"
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 4: Description (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = description
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.border = thin_border
    ws.row_dimensions[current_row].height = 75
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 5: RS and Amount
    ws[f'A{current_row}'] = "RS"
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells(f'B{current_row}:H{current_row}')
    cell = ws[f'B{current_row}']
    cell.value = entry['amount']
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 6: Amount in words + Signatures
    ws.merge_cells(f'A{current_row}:D{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"Rupees {amount_words} only"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{current_row}'].border = thin_border
    
    ws.merge_cells(f'E{current_row}:F{current_row}')
    cell = ws[f'E{current_row}']
    cell.value = "Signature of Payee"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['E', 'F']:
        ws[f'{col}{current_row}'].border = thin_border
    
    ws.merge_cells(f'G{current_row}:H{current_row}')
    cell = ws[f'G{current_row}']
    cell.value = "Signature of witness"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 7: Labour Engaged (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "1. Labour Engaged is Justified"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 8: Work done (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "2.Work is done satisfactorily"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 9: Provision Exists (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "3.Provision Exists in the estimate Maintainnace Grant"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 10: Account code (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = "RM Cables / TMR/LABOUR/5020819"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 11: Passed and Paid (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"Passed and Paid for Rs. {entry['amount']}/-"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Row 12: Amount in words parentheses (merged A-H)
    ws.merge_cells(f'A{current_row}:H{current_row}')
    cell = ws[f'A{current_row}']
    cell.value = f"(Rupees {amount_words.title()} only)"
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{current_row}'].border = thin_border
    current_row += 1
    
    # Add spacing
    current_row += 2
    voucher_no += 1

# Save workbook with dynamic name based on input file
output_filename = f'{base_name}_cash_receipt.xlsx'
wb_new.save(output_filename)
print(f"\n✓ Successfully generated {output_filename}")
print(f"  Total receipts created: {len(ty_data)}")
print(f"  Format: Matched with merged cells and borders")
print(f"  Location: c:\\01.Myuse\\BSNL\\{output_filename}")

wb_source.close()
wb_new.close()
